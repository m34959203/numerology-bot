"""Выгрузка текстов-трактовок из Excel в core/content/data/texts/*.json.

Использование:
    pip install -e ".[excel]"
    python scripts/extract_texts.py --src "ПРОГРАММА МАТРИЦА на русском.xlsx"

Источник: лист «текст» — таблица «ключ → трактовка». Каждый раздел отчёта
(числовые листы 1–36) делает VLOOKUP значения из РАСЧЕТ в свой диапазон листа
«текст». Реестр диапазонов (REGISTRY) восстановлен по формулам числовых листов.

Для разделов с несколькими колонками-парами (ЧЖП, число души) значение по ключу —
список абзацев (по одному на аспект).
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

# topic_slug -> {"src": РАСЧЕТ-источник, "ranges": [диапазоны на листе «текст»]}.
# Несколько диапазонов => мультиаспектный раздел (значение = список абзацев).
REGISTRY: dict[str, dict] = {
    # --- Психоматрица (квадрат Пифагора), ключ = значение клетки или "нет" ---
    "character": {"src": "K32", "ranges": ["B3:C10"]},
    "energy": {"src": "K36", "ranges": ["B13:C19"]},
    "interest": {"src": "K40", "ranges": ["B22:C27"]},
    "health": {"src": "O32", "ranges": ["B30:C34"]},
    "logic": {"src": "O36", "ranges": ["B37:C41"]},
    "labor": {"src": "O40", "ranges": ["B44:C48"]},
    "luck": {"src": "S32", "ranges": ["B51:C55"]},
    "duty": {"src": "S36", "ranges": ["B58:C62"]},
    "memory": {"src": "S40", "ranges": ["B65:C70"]},
    "self_esteem": {"src": "K44", "ranges": ["B73:C82"]},
    "attitude_to_labor": {"src": "O44", "ranges": ["B85:C91"]},
    "purposefulness": {"src": "W32", "ranges": ["B94:C103"]},
    "family_quality": {"src": "W36", "ranges": ["B106:C115"]},
    "stability": {"src": "W40", "ranges": ["B118:C127"]},
    "temperament": {"src": "W28", "ranges": ["B130:C139"]},
    # доп.блоки, ключ = значение «характер» (K32)
    "self_esteem_by_character": {"src": "K32", "ranges": ["E73:F78"]},
    "purposefulness_by_character": {"src": "K32", "ranges": ["E94:F99"]},
    # --- Скаляры (Этап 2) ---
    "behavior_code": {"src": "AR38", "ranges": ["E272:F283"]},  # ключ = строка кода
    "consciousness_level": {"src": "AR40", "ranges": ["B288:C346"]},
    # мультиаспектные: ЧЖП (7 аспектов) и число души (11 аспектов)
    "life_path": {
        "src": "AR34",
        "ranges": [
            "B225:C236",
            "E225:F236",
            "H225:I236",
            "K225:L236",
            "N225:O236",
            "Q225:R236",
            "T225:U236",
        ],
    },
    "soul_number": {
        "src": "AR36",
        "ranges": [
            "B239:C269",
            "E239:F269",
            "H239:I269",
            "K239:L269",
            "N239:O269",
            "Q239:R269",
            "T239:U269",
            "W239:X269",
            "AA239:AB269",
            "AD239:AE269",
            "AG239:AH269",
        ],
    },
    # --- Зависят от Matr (Этап 3+): текст извлекаем заранее, значение — позже ---
    "spiritual_material_balance": {"src": "W44", "ranges": ["B142:C151"]},
    "human_code": {"src": "AR30", "ranges": ["B155:C210"]},  # лист 18 (коды 44/22)
    "personal_year": {"src": "P55", "ranges": ["B350:C358"]},  # листы 24–28
    # --- Прогноз по годам (ключ = расчётное значение года) ---
    # «График кода жизни»: цифра цикла жизненных сил 0–9 → описание периода.
    "life_force_graph": {"src": "Matr!AH", "ranges": ["B213:C222"]},
    # Луна года (B), Солнце года (E, ключ 0 = «нулевой период»), ИТОГ Солнце−Луна (H).
    "moon_year": {"src": "Matr", "ranges": ["B363:C381"]},
    "sun_year": {"src": "Matr", "ranges": ["E363:F382"]},
    "year_total": {"src": "Matr", "ranges": ["H363:I401"]},
}

_CELL = re.compile(r"([A-Z]+)(\d+)")


def _col_to_idx(col: str) -> int:
    idx = 0
    for ch in col:
        idx = idx * 26 + (ord(ch) - 64)
    return idx


def _parse_range(rng: str) -> tuple[int, int, int, int]:
    a, b = rng.split(":")
    m1, m2 = _CELL.match(a), _CELL.match(b)
    return _col_to_idx(m1[1]), _col_to_idx(m2[1]), int(m1[2]), int(m2[2])


def _norm_key(v) -> str:
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _col_values(ws, letter: str, r1: int, r2: int) -> list:
    idx = _col_to_idx(letter)
    rows = list(ws.iter_rows(min_row=r1, max_row=r2, min_col=idx, max_col=idx))
    return [row[0].value for row in rows]


def _kv_table(ws, key_col: str, text_col: str, r1: int, r2: int) -> dict:
    keys = _col_values(ws, key_col, r1, r2)
    texts = _col_values(ws, text_col, r1, r2)
    table = {}
    for k, t in zip(keys, texts, strict=True):
        if k is None or t is None:
            continue
        table[_norm_key(k)] = str(t).strip()
    return table


def extract_moon_sun(wb, texts_dir: Path) -> None:
    """Выгрузить тексты блока «Луна и Солнце по месяцам» и ЧПГ/ЧПМ/ЧПД.

    Источники: лист «Lists луна и солнце» (G:H — код месяца J → текст) и лист «ЧПМ»
    (N:O — ЧПМ 1–9, W:X — ЧПД 1–9, Q:R — заголовок комбо «ЧПМ − ЧПГ», T:U — текст комбо).
    Ключ комбо использует знак U+2212 (минус), не дефис.
    """
    lists = wb["Lists луна и солнце"]
    chpm = wb["ЧПМ"]
    tables = {
        "moon_sun_monthly": _kv_table(lists, "G", "H", 2, 40),
        "personal_month": _kv_table(chpm, "N", "O", 2, 10),
        "personal_day": _kv_table(chpm, "W", "X", 2, 10),
        "personal_combo_title": _kv_table(chpm, "Q", "R", 2, 82),
        "personal_combo_text": _kv_table(chpm, "T", "U", 2, 82),
    }
    for name, table in tables.items():
        (texts_dir / f"{name}.json").write_text(
            json.dumps(table, ensure_ascii=False, indent=1, sort_keys=True), encoding="utf-8"
        )
        print(f"  {name:24} ключей={len(table):3} -> {name}.json")


def extract(src: Path, out: Path) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(src, data_only=True, read_only=True)
    ts = wb["текст"]
    rows = list(ts.iter_rows(max_row=410, max_col=40))

    def cell(r: int, c: int):
        rr = rows[r - 1]
        return rr[c - 1].value if c - 1 < len(rr) else None

    texts_dir = out / "texts"
    texts_dir.mkdir(parents=True, exist_ok=True)
    registry_out: dict[str, dict] = {}

    for topic, spec in REGISTRY.items():
        multi = len(spec["ranges"]) > 1
        table: dict[str, object] = {}
        for rng in spec["ranges"]:
            kc, tc, r1, r2 = _parse_range(rng)
            for r in range(r1, r2 + 1):
                k, t = cell(r, kc), cell(r, tc)
                if k is None or t is None:
                    continue
                key = _norm_key(k)
                text = str(t).strip()
                if multi:
                    table.setdefault(key, []).append(text)
                else:
                    table[key] = text
        path = texts_dir / f"{topic}.json"
        path.write_text(
            json.dumps(table, ensure_ascii=False, indent=1, sort_keys=True), encoding="utf-8"
        )
        registry_out[topic] = {"src": spec["src"], "multi": multi, "keys": len(table)}
        print(f"  {topic:30} ключей={len(table):3} -> {path.name}")

    extract_moon_sun(wb, texts_dir)

    (out / "texts_registry.json").write_text(
        json.dumps(registry_out, ensure_ascii=False, indent=1, sort_keys=True), encoding="utf-8"
    )
    print(f"Реестр: {out / 'texts_registry.json'}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Выгрузка трактовок из Excel в JSON")
    parser.add_argument("--src", required=True, type=Path, help="путь к .xlsx")
    parser.add_argument(
        "--out", default=Path("core/content/data"), type=Path, help="каталог для JSON"
    )
    args = parser.parse_args()
    extract(args.src, args.out)


if __name__ == "__main__":
    main()
